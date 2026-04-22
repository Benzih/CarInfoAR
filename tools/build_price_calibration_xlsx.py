"""
Build an Excel calibration sheet from the app's scan_history.json.

For each scanned car:
  - extract identifying fields + pricing-relevant inputs
  - replicate PriceEstimator.kt logic in Python
  - record every multiplicative factor for easy inspection
  - leave an empty column for the user to fill in the actual Levi-Yitzhak price
  - auto-compute delta % against the current algorithm

Output: price_calibration.xlsx
"""
from __future__ import annotations

import json
import os
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
HISTORY_JSON = ROOT / "scan_history.json"
OUTPUT_XLSX = ROOT / "price_calibration.xlsx"
TODAY = date.today()


# ---------- PriceEstimator.kt — Python port ----------

def parse_date(s):
    if not s:
        return None
    clean = str(s).replace("/", "-").strip()[:10]
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", clean)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    m = re.match(r"^(\d{4})-(\d{2})$", clean[:7])
    if m:
        y, mo = map(int, m.groups())
        try:
            return date(y, mo, 15)
        except ValueError:
            return None
    return None


def age_years(info):
    reg = parse_date(info.get("onRoadDate"))
    if reg is None and info.get("year"):
        reg = date(int(info["year"]), 6, 1)
    if reg is None:
        return None
    days = (TODAY - reg).days
    return max(0.0, days / 365.25)


def base_price(info):
    p = info.get("priceAtRegistration")
    if p and p > 0:
        return float(p)
    p = info.get("catalogPrice")
    if p and p > 0:
        return float(p)
    return None


def currency_for(country):
    return {"NL": "EUR", "GB": "GBP"}.get(country, "ILS")


def age_fuel_factor(age, fuel, country):
    if country == "IL":
        if age <= 1.0:
            yr = 0.87 - age * 0.04
        elif age <= 3.0:
            yr = 0.83 * 0.93 ** (age - 1.0)
        elif age <= 6.0:
            yr = 0.83 * 0.93 ** 2 * 0.92 ** (age - 3.0)
        else:
            yr = 0.83 * 0.93 ** 2 * 0.92 ** 3 * 0.90 ** (age - 6.0)
    else:
        if age <= 1.0:
            yr = 0.82 - age * 0.06
        elif age <= 3.0:
            yr = 0.76 * 0.88 ** (age - 1.0)
        else:
            yr = 0.76 * 0.88 ** 2 * 0.92 ** (age - 3.0)
    f = (fuel or "").lower()
    if "חשמל" in f or "electric" in f or "elektr" in f:
        fm = 0.75 if age <= 1 else (0.85 if age <= 3 else 0.95)
    elif "hybrid" in f or "היבר" in f or "hybride" in f:
        fm = 1.02
    elif "דיזל" in f or "diesel" in f:
        fm = 0.92 if age >= 5 else 0.98
    elif "lpg" in f or "גפ\"מ" in f or "גפמ" in f:
        fm = 0.90
    else:
        fm = 1.0
    return yr * fm


def owners_factor(info):
    o = (info.get("ownership") or "").lower()
    history = info.get("ownershipHistory") or []
    if "מונית" in o or "taxi" in o:
        usage = 0.30
    elif "לימוד" in o or "driving school" in o:
        usage = 0.25
    elif any(k in o for k in ["השכר", "החכר", "חכיר", "ליסינג", "rental", "lease"]):
        usage = 0.18
    elif "חברה" in o or "company" in o:
        usage = 0.12
    elif "ממשלתי" in o or "government" in o or "מוניציפלי" in o:
        usage = 0.25
    else:
        usage = 0.0

    def hist_has(kws):
        return any(any(k in (rec.get("type") or "").lower() for k in kws) for rec in history)

    hist_pen = 0.12 if usage == 0.0 and hist_has(
        ["השכר", "החכר", "חכיר", "ליסינג", "rental", "lease", "מונית", "taxi"]
    ) else 0.0

    hand_count = max(len(history), 1)
    hand_pen = {1: 0.0, 2: 0.05, 3: 0.10, 4: 0.14}.get(hand_count, 0.18)

    penalties = sorted([p for p in (usage, hist_pen, hand_pen) if p > 0], reverse=True)
    factor = 1.0
    weights = [1.0, 0.5, 0.25]
    for i, p in enumerate(penalties):
        factor *= 1.0 - p * (weights[i] if i < len(weights) else 0.25)
    return factor, hand_count, usage, hist_pen, hand_pen


def mileage_factor(km, age, country):
    if not km or km <= 0 or age <= 0:
        return 1.0
    avg = 13_000 if country == "NL" else 15_000
    expected = avg * age
    delta = km - expected
    adj = -0.02 * (delta / 10_000.0)
    return 1.0 + max(-0.20, min(0.12, adj))


def body_factor(body):
    b = (body or "").lower()
    if any(k in b for k in ["suv", "crossover", "פנאי", "ג'יפ", "jeep"]):
        return 1.05
    if "sedan" in b or "סדאן" in b:
        return 1.00
    if "hatchback" in b or "האצ" in b:
        return 0.98
    if any(k in b for k in ["mpv", "minivan", "מיניואן", "van"]):
        return 0.94
    if any(k in b for k in ["coupe", "cabrio", "convertible", "קברי"]):
        return 0.97
    return 1.00


CHINESE = {"BYD","CHERY","GEELY","MG","JAECOO","OMODA","GREAT WALL","HAVAL","NIO",
           "XPENG","LEAPMOTOR","DONGFENG","MAXUS","ZEEKR","SAIC","ROEWE",
           "צ'רי","ביי די","ג'יקו","ליפמוטור","אומודה"}
PREMIUM_REL = {"TOYOTA","LEXUS","HONDA","MAZDA","SUBARU",
               "טויוטה","לקסוס","הונדה","מאזדה","סובארו"}
KOREAN = {"HYUNDAI","KIA","GENESIS","יונדאי","קיה","גנסיס"}
GERMAN = {"BMW","MERCEDES","MERCEDES-BENZ","AUDI","PORSCHE",
          "ב.מ.וו","ב.מ.ו","מרצדס","אאודי","פורשה"}
WEAK = {"FIAT","ALFA","ALFA ROMEO","RENAULT","CITROEN","PEUGEOT","DACIA",
        "פיאט","אלפא","רנו","סיטרואן","פיג'ו","דאצ'יה","דאציה"}


def brand_factor(make, age, country):
    m = (make or "").upper().strip()
    if country == "IL" and any(b in m for b in CHINESE):
        return 0.85 if age <= 1 else (0.78 if age <= 3 else 0.72), "Chinese-IL"
    if any(b in m for b in PREMIUM_REL):
        return 1.06, "Premium-reliable"
    if country == "IL" and any(b in m for b in KOREAN):
        return 1.04, "Korean-IL"
    if any(b in m for b in GERMAN):
        return (0.95 if age >= 3 else 1.00), "Premium-German"
    if any(b in m for b in WEAK):
        return 0.93, "Weak-resale"
    return 1.00, "Standard"


def trim_factor(trim):
    t = (trim or "").lower()
    if any(k in t for k in ["luxury","premium","יוקרה","עליון","top","executive",
                             "inspire","prestige","limited","gls"]):
        return 1.03
    if any(k in t for k in ["base","basic","בסיס","standard"]):
        return 0.98
    return 1.00


def safety_factor(score):
    s = score or 0
    if 7 <= s <= 8: return 1.02
    if 4 <= s <= 6: return 1.00
    if 1 <= s <= 3: return 0.96
    return 1.00


def test_expiry_factor(info):
    d = parse_date(info.get("testValidUntil"))
    if d is None:
        return 1.0
    return 0.94 if d < TODAY else 1.0


def emission_factor(info):
    fec = (info.get("fuelEfficiencyClass") or "").strip().upper()
    if fec:
        return {"A":1.02,"B":1.01,"F":0.95,"G":0.95}.get(fec, 1.00)
    g = info.get("greenIndex")
    if g is not None and 1 <= g <= 15:
        if g <= 3: return 1.02
        if g >= 10: return 0.96
        return 1.00
    co2 = info.get("co2Emissions")
    if co2 is not None:
        if co2 <= 100: return 1.02
        if co2 >= 200: return 0.95
        return 1.00
    return 1.0


def originality_factor(info):
    f = 1.0
    o = (info.get("originality") or "").lower()
    if "לא" in o or "not" in o:
        f *= 0.85
    if info.get("colorChanged"): f *= 0.92
    if info.get("tiresChanged"): f *= 0.98
    return f


def parallel(info):
    imp = (info.get("importerName") or "").lower()
    return "מקביל" in imp or "parallel" in imp


def round_clean(v):
    if v < 10_000: return int(v // 100) * 100
    if v < 100_000: return int(v // 500) * 500
    return int(v // 1000) * 1000


def confidence(info):
    c = 0.5
    if info.get("priceAtRegistration") or info.get("catalogPrice"): c += 0.15
    if info.get("lastTestKm"): c += 0.10
    if info.get("ownership"): c += 0.05
    if info.get("ownershipHistory"): c += 0.05
    if info.get("onRoadDate"): c += 0.10
    if info.get("bodyType"): c += 0.05
    return min(c, 1.0)


def estimate(info):
    base = base_price(info)
    age = age_years(info)
    if base is None or age is None:
        return None

    country = info.get("country") or "IL"
    fuel = info.get("fuelType")

    f_agefuel = age_fuel_factor(age, fuel, country)
    f_own, hand_count, usage_pen, hist_pen, hand_pen = owners_factor(info)
    f_mile = mileage_factor(info.get("lastTestKm"), age, country)
    f_body = body_factor(info.get("bodyType"))
    f_brand, brand_tier = brand_factor(info.get("manufacturer"), age, country)
    f_trim = trim_factor(info.get("trimLevel"))
    f_safety = safety_factor(info.get("safetyScore") or info.get("safetyRating"))
    f_recall = 0.94 if info.get("hasOpenRecall") else 1.0
    f_test = test_expiry_factor(info)
    f_emiss = emission_factor(info)
    f_orig = originality_factor(info)
    f_lpg = 0.88 if info.get("lpgAdded") else 1.0
    f_par = 0.97 if parallel(info) else 1.0
    f_taxi = 0.70 if info.get("isTaxi") else 1.0
    f_exp = 0.85 if info.get("isExported") else 1.0

    factor = (f_agefuel * f_own * f_mile * f_body * f_brand * f_trim *
              f_safety * f_recall * f_test * f_emiss * f_orig * f_lpg *
              f_par * f_taxi * f_exp)

    mid = base * factor
    conf = confidence(info)
    spread = 0.12 + (1.0 - conf) * 0.08

    return {
        "base": base, "age": age, "country": country, "currency": currency_for(country),
        "brand_tier": brand_tier, "hand_count": hand_count,
        "usage_pen": usage_pen, "hist_pen": hist_pen, "hand_pen": hand_pen,
        "f_agefuel": f_agefuel, "f_own": f_own, "f_mile": f_mile, "f_body": f_body,
        "f_brand": f_brand, "f_trim": f_trim, "f_safety": f_safety,
        "f_recall": f_recall, "f_test": f_test, "f_emiss": f_emiss,
        "f_orig": f_orig, "f_lpg": f_lpg, "f_par": f_par, "f_taxi": f_taxi,
        "f_exp": f_exp, "factor": factor,
        "mid": round_clean(mid),
        "low": round_clean(mid * (1 - spread)),
        "high": round_clean(mid * (1 + spread)),
        "confidence": conf,
    }


# ---------- Build the workbook ----------

def build():
    data = json.loads(HISTORY_JSON.read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Scans"
    ws.sheet_view.rightToLeft = True  # Hebrew-friendly

    columns = [
        # ID
        ("plateNumber",       "מספר רכב",          12),
        ("manufacturer",      "יצרן",               12),
        ("model",             "דגם",                18),
        ("year",              "שנה",                7),
        ("trimLevel",         "גימור",              18),
        ("country",           "מדינה",              7),
        # Pricing inputs
        ("priceAtRegistration","מחיר קטלוג (₪)",    14),
        ("lastTestKm",        'ק"מ',                11),
        ("age",               "גיל (שנים)",         10),
        ("ownership",         "בעלות נוכחית",       16),
        ("hand_count",        "יד",                 5),
        ("ownershipHistoryStr","היסטוריית בעלות",   35),
        ("bodyType",          "סוג מרכב",          14),
        ("brand_tier",        "שיוך יצרן",         15),
        ("fuelType",          "דלק",                10),
        ("color",             "צבע",                12),
        ("onRoadDate",        "תאריך עלייה לכביש",  16),
        ("importerName",      "יבואן",              25),
        ("originality",       "מקוריות",            10),
        ("colorChanged",      "החלפת צבע",          10),
        ("tiresChanged",      "החלפת צמיגים",       12),
        ("lpgAdded",          "גפ\"מ",              7),
        ("safetyScore",       "ציון בטיחות",        10),
        ("greenIndex",        "מדד ירוק",           9),
        ("emissionGroup",     "קבוצת זיהום",        10),
        # Factors (breakdown)
        ("f_agefuel",         "f_גיל×דלק",         10),
        ("f_own",             "f_בעלות",           10),
        ("f_mile",            'f_ק"מ',             10),
        ("f_body",             "f_מרכב",           10),
        ("f_brand",           "f_יצרן",            10),
        ("f_trim",            "f_גימור",           10),
        ("f_safety",          "f_בטיחות",          10),
        ("f_recall",          "f_recall",          9),
        ("f_test",            "f_טסט",             9),
        ("f_emiss",           "f_זיהום",           9),
        ("f_orig",            "f_מקוריות",         10),
        ("f_lpg",             "f_גפ\"מ",           9),
        ("f_par",             "f_מקביל",           9),
        ("f_taxi",            "f_מונית",           9),
        ("f_exp",             "f_יצוא",            9),
        ("factor",            "פקטור כולל",        10),
        # Estimates
        ("low",               "הערכה - נמוך",      12),
        ("mid",               "הערכה - אמצע",      12),
        ("high",              "הערכה - גבוה",      12),
        ("confidence",        "ביטחון",             9),
        # User-filled + delta
        ("actualPrice",       "מחיר לוי יצחק (₪)", 16),
        ("deltaPct",          "סטיה %",             10),
        ("notes",             "הערות",              30),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    estimate_fill = PatternFill("solid", fgColor="D9E1F2")
    user_fill = PatternFill("solid", fgColor="FFF2CC")
    delta_fill = PatternFill("solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center", vertical="center")

    # Header row
    for col_idx, (_, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"

    # Data rows — sort newest-first-like: by year desc, then plate
    sorted_data = sorted(
        data,
        key=lambda r: (-(r.get("timestamp") or 0), r.get("plateNumber") or "")
    )

    key_to_col = {k: i + 1 for i, (k, _, _) in enumerate(columns)}

    for row_idx, info in enumerate(sorted_data, start=2):
        est = estimate(info) or {}

        def val(k):
            if k == "age": return round(est.get("age"), 2) if est.get("age") is not None else None
            if k == "hand_count": return est.get("hand_count") or len(info.get("ownershipHistory") or []) or 1
            if k == "brand_tier": return est.get("brand_tier")
            if k == "ownershipHistoryStr":
                hist = info.get("ownershipHistory") or []
                return " ← ".join(
                    f"{(h.get('type') or '').strip()} ({(h.get('date') or '')[:10]})"
                    for h in hist
                ) or None
            if k in ("colorChanged","tiresChanged","lpgAdded"):
                v = info.get(k)
                if v is None: return ""
                return "כן" if v else "לא"
            if k.startswith("f_") or k == "factor":
                v = est.get(k)
                return round(v, 3) if v is not None else None
            if k in ("low","mid","high"):
                return est.get(k)
            if k == "confidence":
                return round(est.get("confidence"), 2) if est.get("confidence") is not None else None
            if k in ("actualPrice","deltaPct","notes"):
                return None
            return info.get(k)

        for k, _, _ in columns:
            col = key_to_col[k]
            cell = ws.cell(row=row_idx, column=col, value=val(k))
            if k in ("low","mid","high","confidence","factor") or k.startswith("f_"):
                cell.fill = estimate_fill
            elif k in ("actualPrice","notes"):
                cell.fill = user_fill
            elif k == "deltaPct":
                cell.fill = delta_fill

        # Delta formula: (actual - mid) / mid
        mid_col = get_column_letter(key_to_col["mid"])
        actual_col = get_column_letter(key_to_col["actualPrice"])
        delta_col = get_column_letter(key_to_col["deltaPct"])
        ws[f"{delta_col}{row_idx}"] = (
            f'=IF(AND(ISNUMBER({actual_col}{row_idx}),ISNUMBER({mid_col}{row_idx}),{mid_col}{row_idx}<>0),'
            f'({actual_col}{row_idx}-{mid_col}{row_idx})/{mid_col}{row_idx},"")'
        )
        ws[f"{delta_col}{row_idx}"].number_format = "0.0%"
        ws[f"{delta_col}{row_idx}"].fill = delta_fill

        # Number formats
        for k in ("priceAtRegistration","lastTestKm","low","mid","high","actualPrice"):
            c = ws.cell(row=row_idx, column=key_to_col[k])
            c.number_format = "#,##0"

    # Summary sheet with aggregate delta stats
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.rightToLeft = True
    last_row = len(sorted_data) + 1
    delta_range = f"Scans!{delta_col}2:{delta_col}{last_row}"
    ws2["A1"] = "מספר רכבים שנסרקו";       ws2["B1"] = len(sorted_data)
    ws2["A2"] = "רכבים עם הערכה";          ws2["B2"] = f"=COUNT(Scans!{mid_col}2:{mid_col}{last_row})"
    ws2["A3"] = "רכבים עם מחיר לוי יצחק";  ws2["B3"] = f"=COUNT(Scans!{actual_col}2:{actual_col}{last_row})"
    ws2["A4"] = "סטיה ממוצעת %";           ws2["B4"] = f"=IFERROR(AVERAGE({delta_range}),\"\")"
    ws2["A5"] = "סטיה חציונית %";          ws2["B5"] = f"=IFERROR(MEDIAN({delta_range}),\"\")"
    ws2["A6"] = "סטיה מוחלטת ממוצעת %";   ws2["B6"] = f"=IFERROR(SUMPRODUCT(ABS({delta_range}))/COUNT({delta_range}),\"\")"
    ws2["A7"] = "סטיה מקסימלית %";         ws2["B7"] = f"=IFERROR(MAX({delta_range}),\"\")"
    ws2["A8"] = "סטיה מינימלית %";         ws2["B8"] = f"=IFERROR(MIN({delta_range}),\"\")"
    for row in range(4, 9):
        ws2.cell(row=row, column=2).number_format = "0.0%"
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    for r in range(1, 9):
        ws2.cell(row=r, column=1).font = Font(bold=True)

    # Instructions
    ws3 = wb.create_sheet("README")
    ws3.sheet_view.rightToLeft = True
    ws3.column_dimensions["A"].width = 110
    lines = [
        "הוראות שימוש",
        "",
        "1. לכל רכב יש שורה עם כל המידע שנסרק (יצרן/דגם/שנה/גימור וכו').",
        "2. עמודות כחולות = פלט של האלגוריתם הנוכחי (פקטורים + הערכת מחיר).",
        "3. עמודה צהובה 'מחיר לוי יצחק' — למלא ידנית את המחיר האמיתי.",
        "4. עמודה ירוקה 'סטיה %' — מתחשבת אוטומטית: (אמת - הערכה) / הערכה.",
        "   - חיובי = האלגוריתם שלנו מעריך נמוך מדי.",
        "   - שלילי = האלגוריתם מעריך גבוה מדי.",
        "5. גיליון 'Summary' מסכם סטיות ממוצעות/חציוניות אחרי שממלאים מחירים.",
        "",
        "פקטורים (כל f_ מוכפל יחד עם מחיר הקטלוג):",
        "  f_גיל×דלק    - עקומת דחת ערך לפי שנים + מודיפייר לסוג דלק",
        "  f_בעלות      - יד + סוג בעלות (פרטי/השכרה/חכירה/מונית...)",
        "  f_ק\"מ        - חריגה ממספר ק\"מ צפוי לפי הגיל",
        "  f_מרכב       - SUV/סדאן/האצ'/קברי'...",
        "  f_יצרן       - טיר: סיני/קוריאני/יפני-אמין/גרמני-פרימיום/חלש",
        "  f_גימור      - טופ/בסיס/רגיל",
        "  f_בטיחות     - ציון בטיחות",
        "  f_recall     - recall פתוח (NL)",
        "  f_טסט        - טסט פג תוקף",
        "  f_זיהום      - green index / CO2 / fuel class",
        "  f_מקוריות    - לא מקורי/החלפת צבע/החלפת צמיגים",
        "  f_גפ\"מ       - הוסף גפ\"מ",
        "  f_מקביל      - יבוא מקביל",
        "  f_מונית      - שימוש במונית",
        "  f_יצוא       - רשום ליצוא",
    ]
    for i, ln in enumerate(lines, start=1):
        ws3.cell(row=i, column=1, value=ln)
        if i == 1:
            ws3.cell(row=i, column=1).font = Font(bold=True, size=14)

    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX} — {len(sorted_data)} rows.")


if __name__ == "__main__":
    build()
