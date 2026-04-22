"""Build xlsx for the 13 new scans — uses v2 estimator as current estimate."""
from __future__ import annotations
import json, sys
from datetime import date
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')

sys.path.insert(0, str(Path(__file__).resolve().parent))
from calibrate_v2 import estimate_v2, age_years

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
HISTORY = ROOT / "scan_history_v2.json"
OUTPUT = ROOT / "price_calibration_check_v2.xlsx"


def main():
    data = json.loads(HISTORY.read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Scans"
    ws.sheet_view.rightToLeft = True

    columns = [
        ("plateNumber", "מספר רכב", 12),
        ("manufacturer", "יצרן", 12),
        ("model", "דגם", 18),
        ("year", "שנה", 7),
        ("trimLevel", "גימור", 16),
        ("country", "מדינה", 7),
        ("priceAtRegistration", "מחיר קטלוג (₪)", 14),
        ("lastTestKm", 'ק"מ', 11),
        ("age", "גיל", 7),
        ("ownership", "בעלות", 14),
        ("hand_count", "יד", 5),
        ("ownershipHistoryStr", "היסטוריית בעלות", 35),
        ("bodyType", "מרכב", 12),
        ("brand_tier", "שיוך יצרן", 15),
        ("fuelType", "דלק", 12),
        ("color", "צבע", 10),
        ("onRoadDate", "עלייה לכביש", 14),
        ("importerName", "יבואן", 20),
        ("originality", "מקוריות", 10),
        ("colorChanged", "החלפת צבע", 10),
        ("tiresChanged", "החלפת צמיגים", 12),
        ("lpgAdded", 'גפ"מ', 7),
        ("safetyScore", "בטיחות", 9),
        ("greenIndex", "מדד ירוק", 9),
        ("f_agefuel", "f_גיל×דלק", 10),
        ("f_own", "f_בעלות", 10),
        ("f_mile", 'f_ק"מ', 10),
        ("f_body", "f_מרכב", 10),
        ("f_brand", "f_יצרן", 10),
        ("f_trim", "f_גימור", 10),
        ("factor", "פקטור כולל", 10),
        ("mid_v2", "הערכה v2 (₪)", 14),
        ("yad2", "יד 2 (₪)", 14),
        ("delta_yad2", "סטיה v2/יד2 %", 12),
        ("levi", "לוי יצחק (₪)", 14),
        ("delta_levi", "סטיה v2/לוי %", 12),
        ("notes", "הערות", 30),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    estimate_fill = PatternFill("solid", fgColor="D9E1F2")
    user_fill = PatternFill("solid", fgColor="FFF2CC")
    delta_fill = PatternFill("solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center", vertical="center")

    for i, (_, label, w) in enumerate(columns, 1):
        c = ws.cell(1, i, label)
        c.font = header_font; c.fill = header_fill; c.alignment = center
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"

    key_to_col = {k: i + 1 for i, (k, _, _) in enumerate(columns)}
    data_sorted = sorted(data, key=lambda r: -(r.get("timestamp") or 0))

    for row_idx, info in enumerate(data_sorted, start=2):
        age = age_years(info)
        if age is None: age = 0.0
        info["age"] = age
        est = estimate_v2(info) or {}

        def val(k):
            if k == "hand_count":
                return len(info.get("ownershipHistory") or []) or 1
            if k == "ownershipHistoryStr":
                hist = info.get("ownershipHistory") or []
                return " ← ".join(
                    f"{(h.get('type') or '').strip()} ({(h.get('date') or '')[:10]})"
                    for h in hist) or None
            if k in ("colorChanged", "tiresChanged", "lpgAdded"):
                v = info.get(k)
                if v is None: return ""
                return "כן" if v else "לא"
            if k == "age":
                return round(age, 2)
            if k in ("f_agefuel","f_own","f_mile","f_body","f_brand","f_trim","factor"):
                v = est.get(k); return round(v, 3) if v is not None else None
            if k == "brand_tier":
                return est.get("brand_tier")
            if k == "mid_v2":
                return int(est.get("mid")) if est.get("mid") else None
            if k in ("yad2", "levi", "delta_yad2", "delta_levi", "notes"):
                return None
            return info.get(k)

        for k, _, _ in columns:
            col = key_to_col[k]
            cell = ws.cell(row_idx, col, val(k))
            if k.startswith("f_") or k in ("factor", "mid_v2"):
                cell.fill = estimate_fill
            elif k in ("yad2", "levi", "notes"):
                cell.fill = user_fill
            elif k.startswith("delta_"):
                cell.fill = delta_fill

        mid_col = get_column_letter(key_to_col["mid_v2"])
        yad2_col = get_column_letter(key_to_col["yad2"])
        levi_col = get_column_letter(key_to_col["levi"])
        dy_col = get_column_letter(key_to_col["delta_yad2"])
        dl_col = get_column_letter(key_to_col["delta_levi"])
        ws[f"{dy_col}{row_idx}"] = (
            f'=IF(AND(ISNUMBER({yad2_col}{row_idx}),ISNUMBER({mid_col}{row_idx}),{mid_col}{row_idx}<>0),'
            f'({yad2_col}{row_idx}-{mid_col}{row_idx})/{mid_col}{row_idx},"")')
        ws[f"{dl_col}{row_idx}"] = (
            f'=IF(AND(ISNUMBER({levi_col}{row_idx}),ISNUMBER({mid_col}{row_idx}),{mid_col}{row_idx}<>0),'
            f'({levi_col}{row_idx}-{mid_col}{row_idx})/{mid_col}{row_idx},"")')
        ws[f"{dy_col}{row_idx}"].number_format = "0.0%"
        ws[f"{dl_col}{row_idx}"].number_format = "0.0%"
        ws[f"{dy_col}{row_idx}"].fill = delta_fill
        ws[f"{dl_col}{row_idx}"].fill = delta_fill
        for k in ("priceAtRegistration", "lastTestKm", "mid_v2", "yad2", "levi"):
            ws.cell(row_idx, key_to_col[k]).number_format = "#,##0"

    # Summary
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.rightToLeft = True
    last = len(data_sorted) + 1
    rng_levi = f"Scans!{dl_col}2:{dl_col}{last}"
    rng_y2 = f"Scans!{dy_col}2:{dy_col}{last}"
    ws2["A1"] = "רכבים שנסרקו"; ws2["B1"] = len(data_sorted)
    ws2["A2"] = "עם לוי יצחק"
    ws2["B2"] = f"=COUNT(Scans!{levi_col}2:{levi_col}{last})"
    ws2["A3"] = "עם יד 2"
    ws2["B3"] = f"=COUNT(Scans!{yad2_col}2:{yad2_col}{last})"
    ws2["A5"] = "סטיה ממוצעת v2 מול לוי";   ws2["B5"] = f"=IFERROR(AVERAGE({rng_levi}),\"\")"
    ws2["A6"] = "סטיה חציונית v2 מול לוי";  ws2["B6"] = f"=IFERROR(MEDIAN({rng_levi}),\"\")"
    ws2["A7"] = "סטיה מוחלטת ממוצעת";      ws2["B7"] = f"=IFERROR(SUMPRODUCT(ABS({rng_levi}))/COUNT({rng_levi}),\"\")"
    ws2["A8"] = "max / min"
    ws2["B8"] = f'=IFERROR(MAX({rng_levi}),"")'; ws2["C8"] = f'=IFERROR(MIN({rng_levi}),"")'
    ws2["A10"] = "סטיה ממוצעת v2 מול יד2";  ws2["B10"] = f"=IFERROR(AVERAGE({rng_y2}),\"\")"
    ws2["A11"] = "MAD v2 מול יד2";          ws2["B11"] = f"=IFERROR(SUMPRODUCT(ABS({rng_y2}))/COUNT({rng_y2}),\"\")"
    for row in (5, 6, 7, 10, 11):
        ws2.cell(row, 2).number_format = "0.0%"
    for r in range(1, 12):
        ws2.cell(r, 1).font = Font(bold=True)
    ws2["B8"].number_format = "0.0%"; ws2["C8"].number_format = "0.0%"
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} — {len(data_sorted)} rows")

    # Also print v2 estimates for sanity
    print("\n=== v2 estimates for the 13 new cars ===")
    print(f"{'plate':<10} {'mfg':<14} {'model':<18} {'yr':<5} {'trim':<14} "
          f"{'cat':<9} {'km':<8} {'age':<5} {'mid_v2':<9} {'ret':<6} {'tier':<16}")
    for info in data_sorted:
        age = age_years(info) or 0
        info["age"] = age
        est = estimate_v2(info)
        if not est:
            print(f"{info['plateNumber']:<10} (no estimate — missing base price or age)")
            continue
        cat = info.get("priceAtRegistration")
        ret = est["mid"] / cat if cat else 0
        print(f"{info['plateNumber']:<10} {str(info.get('manufacturer',''))[:13]:<14} "
              f"{str(info.get('model',''))[:17]:<18} {info.get('year','')!s:<5} "
              f"{str(info.get('trimLevel',''))[:13]:<14} "
              f"{int(cat) if cat else 0:<9,} {str(info.get('lastTestKm',''))[:7]:<8} "
              f"{age:<5.1f} {int(est['mid']):<9,} {ret:<6.2%} {est['brand_tier']:<16}")


if __name__ == "__main__":
    main()
