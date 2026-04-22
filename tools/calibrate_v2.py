"""
Improved PriceEstimator — v2 calibrated against 43 Levi-Yitzhak + 5 Yad2 prices.

Key changes vs v1:
  * AGE CURVE (IL gas): less aggressive Y0-5 (cars hold value), MUCH steeper Y6+
    and Y10+ (old cars drop hard). New data shows Y10+ retention ≈ 0.15-0.20
    vs our old 0.30-0.40 over-estimate.
  * EV/PHEV Y0 PENALTY REMOVED — Chinese EVs in IL are priced to market and
    don't drop 25% in year 1.
  * HYBRID BOOSTED from 1.02 → 1.10 (Toyota/Hyundai hybrids hold value really
    well in IL).
  * DIESEL — staged penalty: mild ≤Y4, steeper Y5-7, very steep Y8+.
  * CHINESE PENALTY SOFTENED Y0-1 (0.85 → 0.93) but tail unchanged.
  * PREMIUM-GERMAN Y5+ STEEPER (0.95 → 0.82): luxury German depreciates hard.
  * HAND PENALTIES HALVED: h4+5+6 were way over-penalizing (data shows old
    cars with many hands still retain normal value).
  * LEASE/RENTAL PENALTY SOFTENED (0.18 → 0.10).
  * MILEAGE SLOPE HALVED and cap softened.
  * FLOOR added: ₪7,000 for IL (scrap-floor observed in LY data).
"""
from __future__ import annotations

import json, math, re, sys
from datetime import date
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
TODAY = date.today()


def parse_date(s):
    if not s: return None
    clean = str(s).replace("/", "-").strip()[:10]
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", clean)
    if m:
        y, mo, d = map(int, m.groups())
        try: return date(y, mo, d)
        except ValueError: return None
    return None


def age_years(info):
    reg = parse_date(info.get("onRoadDate"))
    if reg is None and info.get("year"):
        reg = date(int(info["year"]), 6, 1)
    if reg is None: return None
    days = (TODAY - reg).days
    return max(0.0, days / 365.25)


# ---------- NEW AGE × FUEL CURVE ----------

def age_base_retention(age, country):
    """Pure age-only retention, independent of fuel."""
    if country == "IL":
        if age <= 1.0:
            return 0.92 - 0.07 * age                # 0.92 → 0.85
        if age <= 5.0:
            return 0.85 * 0.925 ** (age - 1.0)      # 0.85 → 0.622
        if age <= 10.0:
            return 0.622 * 0.83 ** (age - 5.0)      # 0.622 → 0.249
        return 0.249 * 0.83 ** (age - 10.0)          # 0.249 → 0.082 @ Y16
    # Non-IL (NL/UK)
    if age <= 1.0:
        return 0.80 - 0.06 * age                     # 0.80 → 0.74
    if age <= 5.0:
        return 0.74 * 0.88 ** (age - 1.0)           # 0.74 → 0.44
    if age <= 10.0:
        return 0.44 * 0.85 ** (age - 5.0)           # 0.44 → 0.195
    return 0.195 * 0.85 ** (age - 10.0)


def fuel_modifier(fuel, age, country, model=None):
    f = (fuel or "").lower()
    m = (model or "").upper()
    # CRITICAL: Israeli data.gov.il often lists Toyota HSD / Hyundai HEV as
    # fuelType="בנזין" — the hybrid nature only shows in the model name.
    hybrid_badge = any(tag in m for tag in ("HSD","HEV"," HYBRID","HYBRID "))
    is_ev = ("חשמל" in f and "בנז" not in f) or ("electric" in f and "hybrid" not in f)
    is_phev = ("חשמל/בנז" in f or "phev" in f or "plug" in f or "plugin" in f
               or ("hybrid" in f and "plug" in f) or "PHEV" in m)
    is_hybrid = (hybrid_badge or "hybrid" in f or "היבר" in f or "hybride" in f) and not is_phev and not is_ev
    is_diesel = "דיזל" in f or "diesel" in f
    is_lpg = "lpg" in f or "גפ\"מ" in f or "גפמ" in f

    if country == "IL":
        if is_ev:
            # EVs in IL: no Y1 penalty (already priced competitively), mild Y2-4, recover later
            if age <= 1.0: return 1.00
            if age <= 3.0: return 0.93
            return 0.96
        if is_phev:
            if age <= 1.0: return 1.02
            if age <= 3.0: return 0.98
            return 1.00
        if is_hybrid:
            # Toyota/Hyundai hybrids hold value unusually well in IL
            # Calibrated against Corolla HSD, Elantra HEV, Ioniq Hybrid data
            return 1.15
        if is_diesel:
            # Diesel passenger cars lose value faster due to ULEZ-type concerns,
            # but commercial vans (Berlingo/Vito/Partner/Caddy/Kangoo/Transit/Doblo)
            # retain value MUCH better because they stay useful as work vehicles.
            commercial_m = (model or "").upper()
            is_commercial_van = any(tag in commercial_m for tag in
                ("BERLINGO","VITO","PARTNER","CADDY","KANGOO","TRANSIT",
                 "DOBLO","DUCATO","SPRINTER","MASTER","TRAFIC","EXPRESS"))
            if is_commercial_van:
                if age <= 4.0: return 0.97
                if age <= 8.0: return 0.98
                return 1.20                          # old commercial van bonus — they stay useful
            if age <= 4.0: return 0.95
            if age <= 7.0: return 0.88
            return 0.80
        if is_lpg:
            return 0.88
        return 1.0  # gasoline

    # Non-IL
    if is_ev:
        if age <= 1.0: return 0.82
        if age <= 3.0: return 0.88
        return 0.95
    if is_hybrid or is_phev:
        return 1.02
    if is_diesel:
        return 0.92 if age >= 5 else 0.98
    if is_lpg:
        return 0.90
    return 1.0


def owners_factor(info):
    o = (info.get("ownership") or "").lower()
    history = info.get("hand_history") or info.get("ownershipHistory") or []
    if isinstance(history, str):
        # When imported from xlsx, ownershipHistory came in as string
        history_items = history.split("←")
    else:
        history_items = [((rec.get("type") or "") if isinstance(rec, dict) else str(rec)) for rec in history]
    hist_text = " ".join(history_items).lower()

    # Usage penalty (biggest single factor)
    if "מונית" in o or "taxi" in o:
        usage = 0.28
    elif "לימוד" in o or "driving school" in o:
        usage = 0.20
    elif any(k in o for k in ["השכר", "החכר", "חכיר", "ליסינג", "rental", "lease"]):
        usage = 0.10
    elif "חברה" in o or "company" in o:
        usage = 0.08
    elif "ממשלתי" in o or "government" in o or "מוניציפלי" in o:
        usage = 0.18
    else:
        usage = 0.0

    hist_fleet = any(k in hist_text for k in ["השכר", "החכר", "חכיר", "ליסינג",
                                              "rental", "lease", "מונית", "taxi"])
    hist_pen = 0.08 if (usage == 0.0 and hist_fleet) else 0.0

    # Hand count
    hand_n = info.get("hand")
    if hand_n is None:
        hand_n = max(len(history_items), 1) if history_items else 1
    hand_n = int(hand_n)
    hand_pen = {1: 0.0, 2: 0.03, 3: 0.06, 4: 0.09, 5: 0.11}.get(hand_n, 0.13)

    penalties = sorted([p for p in (usage, hist_pen, hand_pen) if p > 0], reverse=True)
    factor = 1.0
    weights = [1.0, 0.5, 0.25]
    for i, p in enumerate(penalties):
        w = weights[i] if i < len(weights) else 0.25
        factor *= 1.0 - p * w
    return factor


def mileage_factor(km, age, country):
    if not km or km <= 0 or age <= 0: return 1.0
    avg = 13_000 if country == "NL" else 15_000
    expected = avg * age
    delta = km - expected
    adj = -0.01 * (delta / 10_000.0)  # halved slope
    return 1.0 + max(-0.12, min(0.08, adj))


def body_factor(body):
    b = (body or "").lower()
    if any(k in b for k in ["suv", "crossover", "פנאי", "ג'יפ", "jeep"]): return 1.05
    if "sedan" in b or "סדאן" in b: return 1.00
    if "hatchback" in b or "האצ" in b: return 0.98
    if any(k in b for k in ["mpv", "minivan", "מיניואן", "van"]): return 0.94
    if any(k in b for k in ["coupe", "cabrio", "convertible", "קברי"]): return 0.97
    return 1.00


CHINESE = {"BYD","CHERY","GEELY","MG","JAECOO","OMODA","GREAT WALL","HAVAL","NIO",
           "XPENG","LEAPMOTOR","DONGFENG","MAXUS","ZEEKR","SAIC","ROEWE",
           "צ'רי","ביי די","ג'יקו","ליפמוטור","אומודה","ג'אקו","זיקר","מ.ג"}
PREMIUM_REL = {"TOYOTA","LEXUS","HONDA","MAZDA","SUBARU",
               "טויוטה","לקסוס","הונדה","מאזדה","סובארו"}
KOREAN = {"HYUNDAI","KIA","GENESIS","יונדאי","קיה","גנסיס"}
GERMAN_LUX = {"BMW","MERCEDES","MERCEDES-BENZ","AUDI","PORSCHE",
              "ב.מ.וו","ב.מ.ו","בי אם דבליו","מרצדס","אאודי","פורשה",
              # European luxury that behaves the same way:
              "VOLVO","LAND ROVER","RANGE ROVER","JAGUAR",
              "וולבו","לנד רובר","ריינג רובר","רובר","יגואר"}
WEAK = {"FIAT","ALFA","ALFA ROMEO","RENAULT","CITROEN","PEUGEOT","DACIA","LANCIA",
        "פיאט","אלפא","רנו","סיטרואן","פיג'ו","דאצ'יה","דאציה"}


def brand_factor(make, age, country):
    # IMPORTANT: data.gov.il format is "{brand} {country}" (e.g. "רנו טורקיה").
    # Match as PREFIX — the Hebrew word for Turkey ("טורקיה") ends in "קיה"
    # (Hebrew for Kia), so substring matching would classify Renault Turkey
    # as Korean. Same trap with other Kiev-style cognates.
    m = (make or "").upper().strip()
    def matches(tokens): return any(m.startswith(t) for t in tokens)
    if country == "IL" and matches(CHINESE):
        if age <= 1.0: return 0.93, "Chinese-IL"
        if age <= 3.0: return 0.85, "Chinese-IL"
        return 0.75, "Chinese-IL"
    if matches(PREMIUM_REL):
        return 1.10, "Premium-reliable"
    if country == "IL" and matches(KOREAN):
        return 1.06, "Korean-IL"
    if matches(GERMAN_LUX):
        if age < 3: return 1.00, "Premium-Lux"
        if age < 5: return 0.92, "Premium-Lux"
        return 0.85, "Premium-Lux"
    if matches(WEAK):
        return 0.92, "Weak-resale"
    return 1.00, "Standard"


def trim_factor(trim):
    t = (trim or "").lower()
    top_keywords = ["luxury","premium","יוקרה","עליון","top","executive","inspire",
                    "prestige","limited","gls","supreme","signature","highline",
                    "titanium","business","style"]
    base_keywords = ["base","basic","בסיס","standard","expression","pop","essential"]
    if any(k in t for k in top_keywords): return 1.03
    if any(k in t for k in base_keywords): return 0.97
    return 1.00


def safety_factor(score):
    s = score or 0
    if 7 <= s <= 8: return 1.02
    if 1 <= s <= 3: return 0.96
    return 1.00


def test_expiry_factor(info):
    d = parse_date(info.get("testValidUntil"))
    if d is None: return 1.0
    return 0.94 if d < TODAY else 1.0


def emission_factor(info):
    fec = (info.get("fuelEfficiencyClass") or "").strip().upper()
    if fec: return {"A":1.02,"B":1.01,"F":0.95,"G":0.95}.get(fec, 1.00)
    g = info.get("greenIndex")
    if g is not None and 1 <= g <= 15:
        if g <= 3: return 1.02
        if g >= 10: return 0.96
        return 1.00
    return 1.0


def originality_factor(info):
    f = 1.0
    o = (info.get("originality") or "").lower()
    if "לא" in o or "not" in o: f *= 0.85
    if info.get("colorChanged") in (True, "כן", 1): f *= 0.92
    if info.get("tiresChanged") in (True, "כן", 1): f *= 0.98
    return f


def parallel(info):
    imp = (info.get("importerName") or info.get("importer") or "").lower()
    return "מקביל" in imp or "parallel" in imp


def base_price(info):
    p = info.get("priceAtRegistration") or info.get("price_at_reg")
    if p and p > 0: return float(p)
    p = info.get("catalogPrice")
    if p and p > 0: return float(p)
    return None


def estimate_v2(info):
    base = base_price(info)
    age = info.get("age") if "age" in info and info.get("age") is not None else age_years(info)
    if base is None or age is None: return None
    country = info.get("country") or "IL"
    fuel = info.get("fuelType")

    age_ret = age_base_retention(age, country)
    model = info.get("model")
    fuel_mod = fuel_modifier(fuel, age, country, model=model)
    f_agefuel = age_ret * fuel_mod

    f_own = owners_factor(info)
    f_mile = mileage_factor(info.get("km") or info.get("lastTestKm"), age, country)
    f_body = body_factor(info.get("bodyType"))
    f_brand, brand_tier = brand_factor(info.get("manufacturer") or info.get("mfg"), age, country)
    f_trim = trim_factor(info.get("trimLevel") or info.get("trim"))
    f_safety = safety_factor(info.get("safetyScore") or info.get("safetyRating"))
    f_test = test_expiry_factor(info)
    f_emiss = emission_factor(info)
    f_orig = originality_factor(info)
    f_lpg = 0.88 if (info.get("lpgAdded") in (True, "כן", 1)) else 1.0
    f_par = 0.97 if parallel(info) else 1.0
    f_taxi = 0.70 if info.get("isTaxi") else 1.0
    f_exp = 0.85 if info.get("isExported") else 1.0

    factor = (f_agefuel * f_own * f_mile * f_body * f_brand * f_trim *
              f_safety * f_test * f_emiss * f_orig * f_lpg * f_par * f_taxi * f_exp)
    mid = base * factor

    # IL scrap floor — cars in running condition rarely trade below ~₪8,000
    if country == "IL" and mid < 8000:
        mid = 8000

    return {
        "f_agefuel": f_agefuel, "f_own": f_own, "f_mile": f_mile, "f_body": f_body,
        "f_brand": f_brand, "f_trim": f_trim, "brand_tier": brand_tier,
        "factor": factor, "mid": mid,
    }


# ---------- Validate ----------

def main():
    rows = json.load(open(ROOT / "tools" / "calibration_data.json", encoding="utf-8"))
    results = []
    for r in rows:
        levi = r["levi"] if isinstance(r["levi"], (int, float)) and r["levi"] > 0 else None
        yad2 = r["yad2"] if isinstance(r["yad2"], (int, float)) and r["yad2"] > 0 else None
        target = levi if levi else yad2
        if not target: continue
        v2 = estimate_v2(r)
        if not v2: continue
        v1_mid = r["mid"]
        v2_mid = v2["mid"]
        d1 = (target - v1_mid) / v1_mid
        d2 = (target - v2_mid) / v2_mid
        results.append({
            **r,
            "target": target, "source": "LY" if levi else "Y2",
            "v1_mid": v1_mid, "v2_mid": v2_mid,
            "d1": d1, "d2": d2,
            "improvement": abs(d1) - abs(d2),
        })

    def stats(xs, label):
        xs = sorted(xs)
        if not xs: return
        n = len(xs)
        mean = sum(xs) / n
        med = xs[n // 2]
        mad = sum(abs(x) for x in xs) / n
        print(f"{label}: n={n} mean={mean:+.1%} median={med:+.1%} "
              f"mad={mad:.1%} min={min(xs):+.1%} max={max(xs):+.1%}")

    print("=" * 70)
    print("V1 (current) vs V2 (new) — delta stats vs target price")
    print("=" * 70)
    stats([r["d1"] for r in results], "V1 all")
    stats([r["d2"] for r in results], "V2 all")
    print()
    stats([r["d1"] for r in results if r["source"] == "LY"], "V1 LY only")
    stats([r["d2"] for r in results if r["source"] == "LY"], "V2 LY only")

    print("\n=== 15 worst misses AFTER v2 ===")
    print(f"{'plate':<10} {'mfg':<14} {'model':<18} {'yr':<5} {'age':<5} "
          f"{'cat':<9} {'v1':<8} {'v2':<8} {'tgt':<8} {'d1':<8} {'d2':<8} {'imp':<6}")
    worst = sorted(results, key=lambda p: -abs(p["d2"]))[:15]
    for p in worst:
        print(f"{p['plate']:<10} {str(p['mfg'])[:13]:<14} {str(p['model'])[:17]:<18} "
              f"{str(int(p['year'])):<5} {p['age']:<5.1f} "
              f"{int(p['price_at_reg']):<9,} {int(p['v1_mid']):<8,} "
              f"{int(p['v2_mid']):<8,} {int(p['target']):<8,} "
              f"{p['d1']:+7.1%}  {p['d2']:+7.1%} {p['improvement']:+6.1%}")

    print("\n=== Per-age bucket (v1 vs v2) ===")
    for lo, hi in [(0,2),(2,4),(4,6),(6,10),(10,14),(14,30)]:
        xs = [r for r in results if lo <= r["age"] < hi]
        if not xs: continue
        d1 = [r["d1"] for r in xs]
        d2 = [r["d2"] for r in xs]
        print(f"  age {lo:>2}-{hi:<2}  n={len(xs):<3} "
              f"v1 mean={sum(d1)/len(d1):+6.1%}  v2 mean={sum(d2)/len(d2):+6.1%}   "
              f"v1 mad={sum(abs(x) for x in d1)/len(d1):<6.1%} "
              f"v2 mad={sum(abs(x) for x in d2)/len(d2):<6.1%}")

    print("\n=== Per-brand tier ===")
    for tier in ['Standard','Korean-IL','Premium-reliable','Premium-Lux','Chinese-IL','Weak-resale']:
        xs = [r for r in results if r.get("brand_tier") == tier]
        if not xs: continue
        d1 = [r["d1"] for r in xs]
        d2 = [r["d2"] for r in xs]
        print(f"  {tier:<20} n={len(xs):<2} v1 mean={sum(d1)/len(d1):+6.1%}  "
              f"v2 mean={sum(d2)/len(d2):+6.1%}")

    # Save updated xlsx
    from openpyxl import load_workbook
    wb = load_workbook(r"C:\Users\ASUS\Downloads\סקירקת רכבים למילוי מחירון.xlsx")
    ws = wb["Scans"]
    # Find a free column after col 50 for v2 mid and delta
    ws.cell(1, 51, "הערכה v2 (₪)")
    ws.cell(1, 52, "סטיה v2 vs LY %")
    for row_idx, r in enumerate(results, start=2):
        # Find the right row in xlsx by plate
        for excel_row in range(2, ws.max_row + 1):
            if str(ws.cell(excel_row, 1).value) == str(r["plate"]):
                ws.cell(excel_row, 51, int(r["v2_mid"]))
                ws.cell(excel_row, 52, r["d2"])
                ws.cell(excel_row, 52).number_format = "0.0%"
                break
    out = ROOT / "price_calibration_v2.xlsx"
    wb.save(out)
    print(f"\nSaved {out}")


if __name__ == "__main__":
    main()
