"""
Unified tuning pipeline v3.

Compares multiple PriceEstimator variants against the combined LY calibration set
(43 old cars from calibration_data.json + 12 new from scan_history_v2.json).

Each variant is a pure function VehicleInfo-dict -> estimate-dict so we can bolt
refinements on top and measure the delta.
"""
from __future__ import annotations
import json, sys, math
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent
TODAY = date.today()

# ---------- Shared helpers ----------
def parse_date(s):
    if not s: return None
    c = str(s).replace("/", "-").strip()[:10]
    import re
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", c)
    if m:
        try: return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except: return None
    # Handle MM-YYYY (data.gov.il onRoadDate format)
    m = re.match(r"^(\d{2})-(\d{4})$", c[:7])
    if m:
        try: return date(int(m.group(2)), int(m.group(1)), 15)
        except: return None
    m = re.match(r"^(\d{4})-(\d{2})$", c[:7])
    if m:
        try: return date(int(m.group(1)), int(m.group(2)), 15)
        except: return None
    return None

def age_years(info):
    reg = parse_date(info.get("onRoadDate"))
    if reg is None and info.get("year"):
        reg = date(int(float(info["year"])), 6, 1)
    if reg is None: return None
    return max(0.0, (TODAY - reg).days / 365.25)

def base_price(info):
    p = info.get("priceAtRegistration") or info.get("price_at_reg")
    if p and float(p) > 0: return float(p)
    p = info.get("catalogPrice")
    if p and float(p) > 0: return float(p)
    return None

def round_clean(v):
    if v < 10_000: return int(v // 100) * 100
    if v < 100_000: return int(v // 500) * 500
    return int(v // 1000) * 1000


# ---------- Estimator variants ----------

# Shared brand sets
CHINESE = {"BYD","CHERY","GEELY","MG","JAECOO","OMODA","GREAT WALL","HAVAL","NIO",
    "XPENG","LEAPMOTOR","DONGFENG","MAXUS","ZEEKR","SAIC","ROEWE",
    "צ'רי","ביי די","ג'יקו","ליפמוטור","אומודה","מ.ג","ג'אקו","זיקר"}
PREMIUM_REL_BASE = {"TOYOTA","LEXUS","HONDA","MAZDA","SUBARU",
    "טויוטה","לקסוס","הונדה","מאזדה","סובארו"}
KOREAN = {"HYUNDAI","KIA","GENESIS","יונדאי","קיה","גנסיס"}
GERMAN_LUX = {"BMW","MERCEDES","MERCEDES-BENZ","AUDI","PORSCHE",
    "ב.מ.וו","ב.מ.ו","בי אם דבליו","מרצדס","אאודי","פורשה",
    "VOLVO","LAND ROVER","RANGE ROVER","JAGUAR",
    "וולבו","לנד רובר","ריינג רובר","רובר","יגואר"}
WEAK = {"FIAT","ALFA","ALFA ROMEO","RENAULT","CITROEN","PEUGEOT","DACIA","LANCIA",
    "פיאט","אלפא","רנו","סיטרואן","פיג'ו","דאצ'יה","דאציה"}

def _prefix_match(m, tokens): return any(m.startswith(t) for t in tokens)

def _age_fuel_v2(age, fuel, country):
    """Python v2 curve — explicit anchors."""
    if country == "IL":
        if age <= 1.0:  yr = 0.92 - 0.07 * age
        elif age <= 5.0:  yr = 0.85 * 0.925 ** (age - 1.0)
        elif age <= 10.0: yr = 0.622 * 0.83 ** (age - 5.0)
        else:             yr = 0.249 * 0.83 ** (age - 10.0)
    else:
        if age <= 1.0:  yr = 0.82 - 0.06 * age
        elif age <= 3.0:  yr = 0.76 * 0.88 ** (age - 1.0)
        else:             yr = 0.76 * 0.88**2 * 0.92 ** (age - 3.0)
    f = (fuel or "").lower()
    if "חשמל" in f or "electric" in f:
        fm = 0.75 if age <= 1 else (0.85 if age <= 3 else 0.95)
    elif "hybrid" in f or "היבר" in f or "hybride" in f:
        fm = 1.02
    elif "דיזל" in f or "diesel" in f:
        fm = 0.92 if age >= 5 else 0.98
    else:
        fm = 1.0
    return yr * fm

def _age_fuel_v3(age, fuel, model, country):
    """v3 curve — same base as v2 but with:
    - HYBRID badge detection from MODEL NAME (data.gov.il labels hybrids as 'בנזין')
    - Y10+ curve slightly softer (LY data shows 0.85/yr past Y10, not 0.83)
    """
    if country == "IL":
        if age <= 1.0:  yr = 0.92 - 0.07 * age
        elif age <= 5.0:  yr = 0.85 * 0.925 ** (age - 1.0)
        elif age <= 10.0: yr = 0.622 * 0.83 ** (age - 5.0)
        else:             yr = 0.249 * 0.85 ** (age - 10.0)  # softened past Y10
    else:
        if age <= 1.0:  yr = 0.82 - 0.06 * age
        elif age <= 3.0:  yr = 0.76 * 0.88 ** (age - 1.0)
        else:             yr = 0.76 * 0.88**2 * 0.92 ** (age - 3.0)

    f = (fuel or "").lower()
    mdl = (model or "").upper()
    # Hybrid badge in model name
    if any(tag in mdl for tag in ("HSD","HEV","HYBRID","PHEV","SELF-CHARGING","HYBRYD")):
        fm = 1.08 if country == "IL" else 1.02
    elif "חשמל" in f or "electric" in f:
        fm = 0.75 if age <= 1 else (0.85 if age <= 3 else 0.95)
    elif "hybrid" in f or "היבר" in f or "hybride" in f:
        fm = 1.08 if country == "IL" else 1.02
    elif "דיזל" in f or "diesel" in f:
        fm = 0.92 if age >= 5 else 0.98
    else:
        fm = 1.0
    return yr * fm

def _hand_count_v2(history):
    """v2: counts every entry as a hand."""
    return max(len(history or []), 1)

def _hand_count_v3(history):
    """v3: dealer ('סוחר') entries are not owners — skip them."""
    if not history: return 1
    real = [h for h in history if "סוחר" not in (h.get("type") or "")]
    return max(len(real), 1)

def _owners_factor(info, hand_count_fn):
    o = (info.get("ownership") or "").lower()
    history = info.get("ownershipHistory") or info.get("hand_history") or []
    if isinstance(history, str):
        # legacy "פרטי (2022) ← פרטי (2025)" format
        segs = [s.strip() for s in history.split("←")]
        history = [{"type": s.split("(")[0].strip()} for s in segs if s]
    if "מונית" in o or "taxi" in o: usage = 0.30
    elif "לימוד" in o or "driving school" in o: usage = 0.25
    elif any(k in o for k in ["השכר","החכר","חכיר","ליסינג","rental","lease"]): usage = 0.18
    elif "חברה" in o or "company" in o: usage = 0.12
    elif "ממשלתי" in o or "government" in o or "מוניציפלי" in o: usage = 0.25
    else: usage = 0.0

    def hist_has(kws):
        return any(any(k in (rec.get("type") or "").lower() for k in kws) for rec in history)
    hist_pen = 0.12 if usage == 0.0 and hist_has(["השכר","החכר","חכיר","ליסינג","rental","lease","מונית","taxi"]) else 0.0

    hc = hand_count_fn(history)
    hand_pen = {1: 0.0, 2: 0.05, 3: 0.10, 4: 0.14}.get(hc, 0.18)

    penalties = sorted([p for p in (usage, hist_pen, hand_pen) if p > 0], reverse=True)
    factor = 1.0
    weights = [1.0, 0.5, 0.25]
    for i, p in enumerate(penalties):
        factor *= 1.0 - p * (weights[i] if i < len(weights) else 0.25)
    return factor, hc

def _mile_factor(km, age, country, cap_neg):
    if not km or float(km) <= 0 or age <= 0: return 1.0
    avg = 13_000 if country == "NL" else 15_000
    delta = float(km) - avg * age
    adj = -0.02 * (delta / 10_000.0)
    return 1.0 + max(-cap_neg, min(0.08, adj))

def _mile_factor_v3(km, age, country):
    """v3: cap reduced -0.12 → -0.05 (LY effectively ignores km; don't over-penalize)."""
    return _mile_factor(km, age, country, 0.05)

def _body_factor(body):
    b = (body or "").lower()
    if any(k in b for k in ["suv","crossover","פנאי","ג'יפ","jeep"]): return 1.05
    if "sedan" in b or "סדאן" in b: return 1.00
    if "hatchback" in b or "האצ" in b: return 0.98
    if any(k in b for k in ["mpv","minivan","מיניואן","van"]): return 0.94
    if any(k in b for k in ["coupe","cabrio","convertible","קברי"]): return 0.97
    return 1.00

def _brand_v2(make, age, country, suzuki_tier="standard", korean_mult=1.06):
    """v2 brand classifier with tunable Suzuki and Korean settings."""
    m = (make or "").upper().strip()
    if country == "IL" and _prefix_match(m, CHINESE):
        return (0.93 if age <= 1 else (0.85 if age <= 3 else 0.75)), "Chinese-IL"
    if suzuki_tier == "premium_reliable" and m.startswith("סוזוקי"):
        return 1.08, "Premium-reliable"  # suzuki-boost
    if _prefix_match(m, PREMIUM_REL_BASE):
        return 1.10, "Premium-reliable"
    if country == "IL" and _prefix_match(m, KOREAN):
        return korean_mult, "Korean-IL"
    if _prefix_match(m, GERMAN_LUX):
        if age < 3: return 1.00, "Premium-Lux"
        if age < 5: return 0.92, "Premium-Lux"
        return 0.85, "Premium-Lux"
    if _prefix_match(m, WEAK):
        return 0.92, "Weak-resale"
    return 1.00, "Standard"

def _trim_factor(trim):
    t = (trim or "").lower()
    if any(k in t for k in ["luxury","premium","יוקרה","עליון","top","executive",
                             "inspire","prestige","limited","gls","titanium","signature","supreme"]):
        return 1.03
    if any(k in t for k in ["base","basic","בסיס","standard","pop","expression","essential"]):
        return 0.97
    return 1.00

def _safety(score):
    s = int(score or 0)
    if 7 <= s <= 8: return 1.02
    if 4 <= s <= 6: return 1.00
    if 1 <= s <= 3: return 0.96
    return 1.00

def _emission(info):
    # Only use greenIndex if in valid 1-15 range (some data returns 266)
    g = info.get("greenIndex")
    try: g = float(g) if g is not None else None
    except: g = None
    if g is not None and 1 <= g <= 15:
        return 1.02 if g <= 3 else (0.96 if g >= 10 else 1.00)
    co2 = info.get("co2Emissions")
    if co2:
        try: co2 = float(co2)
        except: co2 = None
        if co2:
            return 1.02 if co2 <= 100 else (0.95 if co2 >= 200 else 1.00)
    return 1.0

def _originality(info):
    f = 1.0
    o = (info.get("originality") or "")
    if "לא" in o or "not" in o.lower(): f *= 0.85
    if info.get("colorChanged") in (True, "כן", 1, "True"): f *= 0.92
    if info.get("tiresChanged") in (True, "כן", 1, "True"): f *= 0.98
    return f


def estimate(info, variant="v3"):
    """Compute estimate under a given variant."""
    base = base_price(info)
    age = age_years(info)
    if base is None or age is None: return None
    country = info.get("country") or "IL"
    fuel = info.get("fuelType")
    model = info.get("model")

    if variant == "kotlin_current":
        # The curve currently live in Kotlin PriceEstimator.kt (as of daa020f)
        if country == "IL":
            if age <= 1.0: yr = 0.87 - 0.04 * age
            elif age <= 3.0: yr = 0.83 * 0.93 ** (age - 1.0)
            elif age <= 6.0: yr = 0.83 * 0.93**2 * 0.92 ** (age - 3.0)
            else: yr = 0.83 * 0.93**2 * 0.92**3 * 0.90 ** (age - 6.0)
        else:
            if age <= 1.0: yr = 0.82 - 0.06 * age
            elif age <= 3.0: yr = 0.76 * 0.88 ** (age - 1.0)
            else: yr = 0.76 * 0.88**2 * 0.92 ** (age - 3.0)
        f = (fuel or "").lower()
        if "חשמל" in f: fm = 0.75 if age <= 1 else (0.85 if age <= 3 else 0.95)
        elif "hybrid" in f or "היבר" in f: fm = 1.02
        elif "דיזל" in f or "diesel" in f: fm = 0.92 if age >= 5 else 0.98
        else: fm = 1.0
        f_agefuel = yr * fm
        f_own, hc = _owners_factor(info, _hand_count_v2)
        f_mile = _mile_factor(info.get("km") or info.get("lastTestKm"), age, country, 0.20)  # old cap
        f_brand, tier = _brand_v2(info.get("mfg") or info.get("manufacturer"), age, country)
    elif variant == "v2_python":
        f_agefuel = _age_fuel_v2(age, fuel, country)
        f_own, hc = _owners_factor(info, _hand_count_v2)
        f_mile = _mile_factor(info.get("km") or info.get("lastTestKm"), age, country, 0.20)
        f_brand, tier = _brand_v2(info.get("mfg") or info.get("manufacturer"), age, country)
    elif variant == "v3":
        f_agefuel = _age_fuel_v3(age, fuel, model, country)
        f_own, hc = _owners_factor(info, _hand_count_v3)
        f_mile = _mile_factor_v3(info.get("km") or info.get("lastTestKm"), age, country)
        f_brand, tier = _brand_v2(info.get("mfg") or info.get("manufacturer"), age, country,
                                    suzuki_tier="premium_reliable", korean_mult=1.08)
    else:
        raise ValueError(variant)

    f_body = _body_factor(info.get("bodyType"))
    f_trim = _trim_factor(info.get("trimLevel") or info.get("trim"))
    f_safety = _safety(info.get("safetyScore"))
    f_emiss = _emission(info)
    f_orig = _originality(info)
    f_lpg = 0.88 if info.get("lpgAdded") in (True,"כן",1) else 1.0

    factor = f_agefuel*f_own*f_mile*f_body*f_brand*f_trim*f_safety*f_emiss*f_orig*f_lpg
    mid = max(base * factor, 8000)
    return {"mid": round_clean(mid), "factor": factor, "tier": tier, "hand": hc,
            "f_agefuel": f_agefuel, "f_own": f_own, "f_mile": f_mile,
            "f_brand": f_brand, "f_body": f_body, "f_trim": f_trim,
            "f_safety": f_safety, "f_emiss": f_emiss, "f_orig": f_orig, "f_lpg": f_lpg,
            "age": age}


# ---------- Load combined dataset ----------

def load_combined():
    """Return list of rows with levi+full info. Combines both sources."""
    rows = []

    # Source A: calibration_data.json (43 LY cars)
    cal = json.load(open(ROOT/"tools/calibration_data.json", encoding='utf-8'))
    for r in cal:
        if not r.get('levi'): continue
        # Synthesize VehicleInfo-ish dict
        info = {
            "plateNumber": r.get("plate"),
            "manufacturer": r.get("mfg"),
            "model": r.get("model"),
            "year": r.get("year"),
            "trimLevel": r.get("trim"),
            "country": r.get("country", "IL"),
            "priceAtRegistration": r.get("price_at_reg"),
            "lastTestKm": r.get("km"),
            "ownership": r.get("ownership"),
            "ownershipHistory": [],  # synthesized below
            "onRoadDate": r.get("onRoadDate"),
            "fuelType": r.get("fuelType"),
            "bodyType": r.get("bodyType"),
            "safetyScore": r.get("safetyScore"),
            "greenIndex": r.get("greenIndex"),
            "colorChanged": r.get("colorChanged"),
            "tiresChanged": r.get("tiresChanged"),
            "lpgAdded": r.get("lpgAdded"),
            "originality": r.get("originality"),
            "importerName": r.get("importer"),
        }
        # Parse hand_history string like "פרטי (08/2022) ← פרטי (02/2025)"
        hh = r.get("hand_history") or ""
        segs = [s.strip() for s in hh.split("←") if s.strip()]
        info["ownershipHistory"] = [{"type": s.split("(")[0].strip(),
                                       "date": s.split("(")[1].rstrip(")") if "(" in s else ""}
                                      for s in segs]
        rows.append({"info": info, "levi": r["levi"], "yad2": r.get("yad2"),
                     "source": "old"})

    # Source B: new scan + filled xlsx (12 LY cars)
    wb = load_workbook(ROOT/"price_calibration_check_v2_with_levieitshak.xlsx", data_only=True)
    ws = wb['Scans']
    # Get plate → levi from xlsx
    plate_to_levi = {}
    plate_to_yad2 = {}
    for row in range(2, ws.max_row+1):
        plate = str(ws.cell(row,1).value or "")
        lv = ws.cell(row,35).value
        y2 = ws.cell(row,33).value
        if plate: plate_to_levi[plate] = lv; plate_to_yad2[plate] = y2

    sh = json.load(open(ROOT/"scan_history_v2.json", encoding='utf-8'))
    for info in sh:
        plate = str(info.get("plateNumber") or "")
        lv = plate_to_levi.get(plate)
        if not (isinstance(lv,(int,float)) and lv > 0):
            continue
        rows.append({"info": info, "levi": lv, "yad2": plate_to_yad2.get(plate),
                     "source": "new"})
    return rows


# ---------- Report ----------

def stats(deltas, label):
    deltas = sorted(deltas)
    n = len(deltas)
    if n == 0: return
    mean = sum(deltas)/n
    med = deltas[n//2]
    mad = sum(abs(d) for d in deltas)/n
    within_10 = sum(1 for d in deltas if abs(d) <= 0.10) / n
    within_20 = sum(1 for d in deltas if abs(d) <= 0.20) / n
    print(f"{label}: n={n}  MAD={mad*100:5.2f}%  mean={mean*100:+5.2f}%  "
          f"med={med*100:+5.2f}%  ±10%={within_10:.0%}  ±20%={within_20:.0%}  "
          f"min={min(deltas)*100:+.1f}%  max={max(deltas)*100:+.1f}%")


def main():
    rows = load_combined()
    print(f"Combined dataset: {len(rows)} LY-priced cars "
          f"({sum(1 for r in rows if r['source']=='old')} old + "
          f"{sum(1 for r in rows if r['source']=='new')} new)\n")

    for variant in ("kotlin_current", "v2_python", "v3"):
        deltas_all = []
        deltas_old = []
        deltas_new = []
        fails = []
        for r in rows:
            e = estimate(r['info'], variant)
            if not e:
                fails.append(r['info'].get('plateNumber')); continue
            d = (r['levi'] - e['mid']) / e['mid']
            deltas_all.append(d)
            (deltas_old if r['source']=='old' else deltas_new).append(d)
        print(f"=== {variant} (fails: {len(fails)}) ===")
        stats(deltas_all, "  ALL ")
        stats(deltas_old, "  old ")
        stats(deltas_new, "  new ")
        print()

    # Per-car table for v3 to inspect individual deltas
    print("=== v3 per-car (sorted by abs delta) ===")
    results = []
    for r in rows:
        e = estimate(r['info'], "v3")
        if not e: continue
        d = (r['levi'] - e['mid']) / e['mid']
        results.append((r, e, d))
    results.sort(key=lambda x: -abs(x[2]))

    print(f"{'plate':<10} {'mfg':<12} {'model':<16} {'yr':<5} {'age':<5} "
          f"{'cat':<9} {'km':<8} {'mid':<9} {'LY':<9} {'delta':<8} {'tier':<14}")
    for r, e, d in results[:25]:
        info = r['info']
        mfg = str(info.get('manufacturer') or info.get('mfg') or '')[:11]
        model = str(info.get('model') or '')[:15]
        cat = base_price(info)
        km = info.get('lastTestKm') or info.get('km')
        print(f"{str(info.get('plateNumber') or ''):<10} "
              f"{mfg:<12} {model:<16} {info.get('year','')!s:<5} "
              f"{e['age']:<5.1f} "
              f"{(str(int(cat)) if cat else '-'):<9} "
              f"{(str(int(km)) if km else '-'):<8} "
              f"{e['mid']:<9} {int(r['levi']):<9} {d*100:+7.1f}% {e['tier']:<14}")


if __name__ == "__main__":
    main()
