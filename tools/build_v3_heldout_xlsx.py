"""Build xlsx for the fresh hold-out test set (scan_history_v3_test.json).

Uses the v3 estimator (identical logic to what's now in Kotlin PriceEstimator.kt).
Leaves empty columns for Levi-Yitzhak + Yad2 for the user to fill.
"""
from __future__ import annotations
import json, sys
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT/"tools"))

from build_v3_validation_xlsx import est_v3
from tune_v3 import base_price, age_years

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HISTORY = ROOT / "scan_history_v3_test.json"
OUTPUT = ROOT / "price_calibration_v3_heldout.xlsx"


def main():
    data = json.loads(HISTORY.read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Held-out"
    ws.sheet_view.rightToLeft = True

    cols = [
        ("plate", "מספר", 11),
        ("mfg", "יצרן", 14),
        ("model", "דגם", 19),
        ("year", "שנה", 6),
        ("trim", "גימור", 14),
        ("age", "גיל", 5),
        ("cat", "מחיר קטלוג", 12),
        ("km", 'ק"מ', 9),
        ("hand", "יד", 5),
        ("ownership", "בעלות", 12),
        ("hand_history", "היסטוריית ידיים", 32),
        ("bodyType", "מרכב", 12),
        ("tier", "שכבה", 16),
        ("fuel", "דלק", 12),
        ("f_af", "f_גיל×דלק", 10),
        ("f_own", "f_בעלות", 10),
        ("f_mile", 'f_ק"מ', 10),
        ("f_brand", "f_יצרן", 10),
        ("factor", "פקטור", 8),
        ("mid_v3", "הערכה v3 (₪)", 13),
        ("yad2", "יד 2 (₪)", 12),
        ("delta_yad2", "Δ v3/יד2", 10),
        ("levi", "לוי יצחק (₪)", 13),
        ("delta_levi", "Δ v3/לוי", 10),
        ("notes", "הערות", 30),
    ]

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    est_fill = PatternFill("solid", fgColor="D9E1F2")
    user_fill = PatternFill("solid", fgColor="FFF2CC")
    delta_fill = PatternFill("solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center", vertical="center")

    for i, (_, lbl, w) in enumerate(cols, 1):
        c = ws.cell(1, i, lbl)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "C2"

    key2col = {k: i + 1 for i, (k, _, _) in enumerate(cols)}
    data_sorted = sorted(data, key=lambda r: -(r.get("timestamp") or 0))

    for ri, info in enumerate(data_sorted, 2):
        e = est_v3(info) or {}

        def v(k):
            if k == "plate": return info.get("plateNumber")
            if k == "mfg": return info.get("manufacturer")
            if k == "model": return info.get("model")
            if k == "year": return info.get("year")
            if k == "trim": return info.get("trimLevel")
            if k == "age": return round(e.get("age", 0), 1) if e else None
            if k == "cat": return base_price(info)
            if k == "km": return info.get("lastTestKm")
            if k == "hand": return e.get("hand") if e else None
            if k == "ownership": return info.get("ownership")
            if k == "hand_history":
                hist = info.get("ownershipHistory") or []
                return " ← ".join(f"{(h.get('type','')).strip()} ({(h.get('date','') or '')[:10]})"
                                    for h in hist) or None
            if k == "bodyType": return info.get("bodyType")
            if k == "tier": return e.get("tier") if e else None
            if k == "fuel": return info.get("fuelType")
            if k in ("f_af", "f_own", "f_mile", "f_brand", "factor"):
                val = e.get(k) if e else None
                return round(val, 3) if val is not None else None
            if k == "mid_v3": return int(e["mid"]) if e.get("mid") else None
            return None  # user-filled columns

        for k, _, _ in cols:
            col = key2col[k]
            cell = ws.cell(ri, col, v(k))
            if k.startswith("f_") or k in ("factor", "mid_v3"):
                cell.fill = est_fill
            elif k in ("yad2", "levi", "notes"):
                cell.fill = user_fill
            elif k.startswith("delta_"):
                cell.fill = delta_fill

        mid_col = get_column_letter(key2col["mid_v3"])
        y2_col = get_column_letter(key2col["yad2"])
        ly_col = get_column_letter(key2col["levi"])
        dy_col = get_column_letter(key2col["delta_yad2"])
        dl_col = get_column_letter(key2col["delta_levi"])
        ws[f"{dy_col}{ri}"] = (
            f'=IF(AND(ISNUMBER({y2_col}{ri}),ISNUMBER({mid_col}{ri}),{mid_col}{ri}<>0),'
            f'({y2_col}{ri}-{mid_col}{ri})/{mid_col}{ri},"")')
        ws[f"{dl_col}{ri}"] = (
            f'=IF(AND(ISNUMBER({ly_col}{ri}),ISNUMBER({mid_col}{ri}),{mid_col}{ri}<>0),'
            f'({ly_col}{ri}-{mid_col}{ri})/{mid_col}{ri},"")')
        ws[f"{dy_col}{ri}"].number_format = "0.0%"
        ws[f"{dl_col}{ri}"].number_format = "0.0%"
        for k in ("cat", "km", "mid_v3", "yad2", "levi"):
            ws.cell(ri, key2col[k]).number_format = "#,##0"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.rightToLeft = True
    n_last = len(data_sorted) + 1
    ws2["A1"] = "מדד";                          ws2["B1"] = "ערך"
    ws2["A2"] = "סה\"כ רכבים";                 ws2["B2"] = len(data_sorted)
    ws2["A3"] = "מולאו עם לוי יצחק";            ws2["B3"] = f"=COUNT('Held-out'!{ly_col}2:{ly_col}{n_last})"
    ws2["A4"] = "מולאו עם יד 2";               ws2["B4"] = f"=COUNT('Held-out'!{y2_col}2:{y2_col}{n_last})"
    ws2["A6"] = "MAD מול לוי";                 ws2["B6"] = f"=IFERROR(SUMPRODUCT(ABS('Held-out'!{dl_col}2:{dl_col}{n_last}))/COUNT('Held-out'!{dl_col}2:{dl_col}{n_last}),\"\")"
    ws2["A7"] = "ממוצע Δ מול לוי";             ws2["B7"] = f"=IFERROR(AVERAGE('Held-out'!{dl_col}2:{dl_col}{n_last}),\"\")"
    ws2["A8"] = "חציון Δ מול לוי";             ws2["B8"] = f"=IFERROR(MEDIAN('Held-out'!{dl_col}2:{dl_col}{n_last}),\"\")"
    ws2["A9"] = "מחזיק ±10% מול לוי";          ws2["B9"] = f"=IFERROR(COUNTIFS('Held-out'!{dl_col}2:{dl_col}{n_last},\">-0.1\",'Held-out'!{dl_col}2:{dl_col}{n_last},\"<0.1\")/COUNT('Held-out'!{dl_col}2:{dl_col}{n_last}),\"\")"
    ws2["A10"] = "מחזיק ±20% מול לוי";         ws2["B10"] = f"=IFERROR(COUNTIFS('Held-out'!{dl_col}2:{dl_col}{n_last},\">-0.2\",'Held-out'!{dl_col}2:{dl_col}{n_last},\"<0.2\")/COUNT('Held-out'!{dl_col}2:{dl_col}{n_last}),\"\")"
    ws2["A11"] = "max Δ";                     ws2["B11"] = f"=IFERROR(MAX('Held-out'!{dl_col}2:{dl_col}{n_last}),\"\")"
    ws2["A12"] = "min Δ";                     ws2["B12"] = f"=IFERROR(MIN('Held-out'!{dl_col}2:{dl_col}{n_last}),\"\")"
    for r in range(6, 13):
        ws2.cell(r, 2).number_format = "0.00%" if r not in (11, 12) else "+0.0%;-0.0%;0.0%"
    for r in range(1, 13):
        ws2.cell(r, 1).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} — {len(data_sorted)} rows")


if __name__ == "__main__":
    main()
