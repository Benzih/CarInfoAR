"""Build validation xlsx: all 55 calibration cars, Kotlin-current vs v3."""
from __future__ import annotations
import sys
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT/"tools"))
from tune_v3 import load_combined, base_price, stats, estimate as est_old
from tune_v6 import estimate_v6   # alias; v7 is inline in last test

# Inline v7 estimator (final)
from tune_v3 import (_body_factor, _trim_factor, _safety, _emission,
    _originality, _owners_factor, _hand_count_v3, _prefix_match,
    round_clean, age_years)
from tune_v4 import is_performance, is_commercial, OLD_GENERIC
from tune_v5 import CHINESE_V5, PREMIUM_REL_V5, KOREAN_V5, GERMAN_LUX_V5, WEAK_V5, MID_RELIABLE_V5

def af_v7(age, fuel, model, country):
    if country == "IL":
        if age <= 1.0:  yr = 0.92 - 0.07 * age
        elif age <= 5.0:  yr = 0.85 * 0.925 ** (age - 1.0)
        elif age <= 10.0: yr = 0.622 * 0.83 ** (age - 5.0)
        else:             yr = 0.249 * 0.85 ** (age - 10.0)
    else:
        if age <= 1.0:  yr = 0.82 - 0.06 * age
        elif age <= 3.0:  yr = 0.76 * 0.88 ** (age - 1.0)
        else:             yr = 0.76 * 0.88**2 * 0.92 ** (age - 3.0)
    f = (fuel or "").lower(); mdl = (model or "").upper()
    is_hyb = any(t in mdl for t in ("HSD","HEV","HYBRID","PHEV","SELF-CHARGING","HYBRYD")) or "hybrid" in f or "היבר" in f
    if is_hyb:
        fm = 1.15 if (country == "IL" and age >= 3.0) else (1.10 if country == "IL" else 1.02)
    elif "חשמל" in f or "electric" in f:
        fm = 0.75 if age <= 1 else (0.85 if age <= 3 else 0.95)
    elif "דיזל" in f or "diesel" in f: fm = 0.92 if age >= 5 else 0.98
    else: fm = 1.0
    return yr * fm

def mile_v7(km, age, country):
    if not km or float(km) <= 0 or age <= 0: return 1.0
    avg = 13000 if country == "NL" else 15000
    adj = -0.01 * ((float(km) - avg*age) / 10000)
    return 1.0 + max(-0.05, min(0.08, adj))

def brand_v7(make, model, age, country):
    m = (make or "").upper().strip(); mdl = (model or "").upper()
    if is_performance(mdl) and _prefix_match(m, GERMAN_LUX_V5):
        if age < 3: return 0.95, "Performance-Lux"
        if age < 5: return 0.82, "Performance-Lux"
        return 0.70, "Performance-Lux"
    if is_commercial(mdl):
        if age < 5: return 1.00, "Commercial"
        return 1.15, "Commercial"
    if country == "IL" and _prefix_match(m, CHINESE_V5):
        if age <= 1.0: return 1.00, "Chinese-IL"
        if age <= 3.0: return 0.95, "Chinese-IL"
        if age <= 5.0: return 0.88, "Chinese-IL"
        return 0.78, "Chinese-IL"
    if _prefix_match(m, PREMIUM_REL_V5):
        if age < 2: return 1.00, "Premium-reliable"
        return 1.10, "Premium-reliable"
    if m.startswith("סוזוקי") or m.startswith("SUZUKI"):
        if age < 2: return 1.00, "Suzuki-solid"
        return 1.05, "Suzuki-solid"
    if country == "IL" and _prefix_match(m, KOREAN_V5): return 1.10, "Korean-IL"
    if _prefix_match(m, GERMAN_LUX_V5):
        if age < 3: return 1.00, "Premium-Lux"
        if age < 5: return 0.92, "Premium-Lux"
        return 0.85, "Premium-Lux"
    if _prefix_match(m, WEAK_V5):
        if age > 10: return 0.78, "Weak-resale"
        return 0.92, "Weak-resale"
    if _prefix_match(m, OLD_GENERIC):
        if age > 10: return 0.70, "Old-generic"
        return 1.00, "Old-generic"
    if _prefix_match(m, MID_RELIABLE_V5): return 1.02, "Mid-reliable"
    return 1.00, "Standard"

def est_v3(info):
    base = base_price(info); age = age_years(info)
    if base is None or age is None: return None
    c = info.get("country") or "IL"
    f_af = af_v7(age, info.get("fuelType"), info.get("model"), c)
    f_own, hc = _owners_factor(info, _hand_count_v3)
    f_mile = mile_v7(info.get("lastTestKm") or info.get("km"), age, c)
    f_body = _body_factor(info.get("bodyType"))
    f_brand, tier = brand_v7(info.get("manufacturer") or info.get("mfg"),
                              info.get("model"), age, c)
    f_trim = _trim_factor(info.get("trimLevel") or info.get("trim"))
    f_safe = _safety(info.get("safetyScore"))
    f_emi = _emission(info)
    f_or = _originality(info)
    f_lpg = 0.88 if info.get("lpgAdded") in (True,"כן",1) else 1.0
    factor = f_af*f_own*f_mile*f_body*f_brand*f_trim*f_safe*f_emi*f_or*f_lpg
    return {"mid": round_clean(max(base*factor,8000)), "tier": tier, "age": age,
            "hand": hc, "f_af": f_af, "f_own": f_own, "f_mile": f_mile,
            "f_brand": f_brand, "factor": factor}


# ---- Build xlsx ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

rows = load_combined()
wb = Workbook()
ws = wb.active
ws.title = "v3 vs LY"
ws.sheet_view.rightToLeft = True

cols = [
    ("plate","מספר",10),("mfg","יצרן",13),("model","דגם",18),("year","שנה",6),
    ("age","גיל",5),("country","מ'",5),("cat","מחיר קטלוג",11),("km",'ק"מ',9),
    ("hand","יד",5),("tier","שכבה",16),
    ("f_af","f_גיל×דלק",9),("f_own","f_בעלות",9),("f_mile",'f_ק"מ',9),
    ("f_brand","f_יצרן",9),("factor","פקטור",8),
    ("mid_old","Kotlin ישן",11),("mid_v3","v3 חדש",11),
    ("levi","לוי יצחק",11),
    ("delta_old","Δ ישן",8),("delta_v3","Δ v3",8),("notes","הערות",30),
]
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F3864")
old_fill = PatternFill("solid", fgColor="F4B084")
new_fill = PatternFill("solid", fgColor="C6EFCE")
ly_fill = PatternFill("solid", fgColor="FFF2CC")
center = Alignment(horizontal="center", vertical="center")

for i, (_, lbl, w) in enumerate(cols, 1):
    c = ws.cell(1, i, lbl); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 32
ws.freeze_panes = "C2"

key2col = {k: i+1 for i,(k,_,_) in enumerate(cols)}
rows_sorted = sorted(rows, key=lambda r: -(r['info'].get('timestamp') or 0))

for ri, r in enumerate(rows_sorted, 2):
    info = r['info']
    old = est_old(info, "kotlin_current")
    new = est_v3(info)
    levi = r['levi']

    def getv(k):
        if k == "plate": return info.get("plateNumber") or info.get("plate")
        if k == "mfg": return info.get("manufacturer") or info.get("mfg")
        if k == "model": return info.get("model")
        if k == "year": return info.get("year")
        if k == "age": return round(new["age"],1) if new else None
        if k == "country": return info.get("country") or "IL"
        if k == "cat": return base_price(info)
        if k == "km": return info.get("lastTestKm") or info.get("km")
        if k == "hand": return new["hand"] if new else None
        if k == "tier": return new["tier"] if new else None
        if k in ("f_af","f_own","f_mile","f_brand","factor"):
            return round(new[k],3) if new else None
        if k == "mid_old": return old["mid"] if old else None
        if k == "mid_v3": return new["mid"] if new else None
        if k == "levi": return levi
        if k == "delta_old":
            if old and levi: return (levi - old["mid"]) / old["mid"]
        if k == "delta_v3":
            if new and levi: return (levi - new["mid"]) / new["mid"]
        return None

    for k, _, _ in cols:
        col = key2col[k]
        cell = ws.cell(ri, col, getv(k))
        if k in ("mid_old","delta_old"): cell.fill = old_fill
        elif k in ("mid_v3","delta_v3"): cell.fill = new_fill
        elif k == "levi": cell.fill = ly_fill
        if k in ("cat","km","mid_old","mid_v3","levi"): cell.number_format = "#,##0"
        if k.startswith("delta"): cell.number_format = "+0.0%;-0.0%;0.0%"

# Summary
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.rightToLeft = True
n = len(rows_sorted) + 1
do_col = get_column_letter(key2col["delta_old"])
dv_col = get_column_letter(key2col["delta_v3"])
mo_col = get_column_letter(key2col["mid_old"])
mv_col = get_column_letter(key2col["mid_v3"])
ly_col = get_column_letter(key2col["levi"])

ws2["A1"] = "מדד";                       ws2["B1"] = "Kotlin ישן"; ws2["C1"] = "v3 חדש"
ws2["A2"] = "MAD (סטיה מוחלטת ממוצעת)";  ws2["B2"] = f"=SUMPRODUCT(ABS('v3 vs LY'!{do_col}2:{do_col}{n}))/COUNT('v3 vs LY'!{do_col}2:{do_col}{n})"; ws2["C2"] = f"=SUMPRODUCT(ABS('v3 vs LY'!{dv_col}2:{dv_col}{n}))/COUNT('v3 vs LY'!{dv_col}2:{dv_col}{n})"
ws2["A3"] = "מחזיק ±10%";                ws2["B3"] = f"=COUNTIFS('v3 vs LY'!{do_col}2:{do_col}{n},\">-0.1\",'v3 vs LY'!{do_col}2:{do_col}{n},\"<0.1\")/COUNT('v3 vs LY'!{do_col}2:{do_col}{n})"; ws2["C3"] = f"=COUNTIFS('v3 vs LY'!{dv_col}2:{dv_col}{n},\">-0.1\",'v3 vs LY'!{dv_col}2:{dv_col}{n},\"<0.1\")/COUNT('v3 vs LY'!{dv_col}2:{dv_col}{n})"
ws2["A4"] = "מחזיק ±20%";                ws2["B4"] = f"=COUNTIFS('v3 vs LY'!{do_col}2:{do_col}{n},\">-0.2\",'v3 vs LY'!{do_col}2:{do_col}{n},\"<0.2\")/COUNT('v3 vs LY'!{do_col}2:{do_col}{n})"; ws2["C4"] = f"=COUNTIFS('v3 vs LY'!{dv_col}2:{dv_col}{n},\">-0.2\",'v3 vs LY'!{dv_col}2:{dv_col}{n},\"<0.2\")/COUNT('v3 vs LY'!{dv_col}2:{dv_col}{n})"
ws2["A5"] = "ממוצע Δ";                   ws2["B5"] = f"=AVERAGE('v3 vs LY'!{do_col}2:{do_col}{n})"; ws2["C5"] = f"=AVERAGE('v3 vs LY'!{dv_col}2:{dv_col}{n})"
ws2["A6"] = "חציון Δ";                   ws2["B6"] = f"=MEDIAN('v3 vs LY'!{do_col}2:{do_col}{n})"; ws2["C6"] = f"=MEDIAN('v3 vs LY'!{dv_col}2:{dv_col}{n})"
ws2["A7"] = "max Δ";                     ws2["B7"] = f"=MAX('v3 vs LY'!{do_col}2:{do_col}{n})"; ws2["C7"] = f"=MAX('v3 vs LY'!{dv_col}2:{dv_col}{n})"
ws2["A8"] = "min Δ";                     ws2["B8"] = f"=MIN('v3 vs LY'!{do_col}2:{do_col}{n})"; ws2["C8"] = f"=MIN('v3 vs LY'!{dv_col}2:{dv_col}{n})"
ws2["A9"] = "סה\"כ רכבים";              ws2["B9"] = f"=COUNT('v3 vs LY'!{do_col}2:{do_col}{n})"; ws2["C9"] = f"=COUNT('v3 vs LY'!{dv_col}2:{dv_col}{n})"

for r in range(1,10):
    for c in range(1,4):
        ws2.cell(r,c).font = Font(bold=r==1)
for r in (2,3,4,5,6,7,8):
    for c in (2,3): ws2.cell(r,c).number_format = "0.00%"
ws2.column_dimensions["A"].width = 32; ws2.column_dimensions["B"].width = 14; ws2.column_dimensions["C"].width = 14

outp = ROOT / "price_calibration_v3_final.xlsx"
wb.save(outp)
print(f"Saved: {outp}")
print(f"Rows: {len(rows_sorted)}")
